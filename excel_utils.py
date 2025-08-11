import openpyxl
from openpyxl.styles import Font
import os
import re

EXCEL_FILE = "recipe.xlsx"

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "材料"
        sheet["A1"] = "材料名"
        sheet["B1"] = "量"
        sheet["C1"] = "価格"
        sheet["D1"] = "税率"
        bold_font = Font(bold=True)
        for col in ["A1", "B1", "C1", "D1"]:
            sheet[col].font = bold_font
        recipe_sheet = wb.create_sheet("レシピ")
        recipe_sheet["A1"] = "レシピ名"
        recipe_sheet["B1"] = "材料名"
        recipe_sheet["C1"] = "使用量ml"
        recipe_sheet["D1"] = "備考"
        for col in ["A1", "B1", "C1", "D1"]:
            recipe_sheet[col].font = bold_font
        wb.save(EXCEL_FILE)

def save_material(name, amount, price, tax):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["材料"]
    sheet.append([name, amount, price, tax])
    wb.save(EXCEL_FILE)

def get_materials():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["材料"]
    return list(sheet.iter_rows(min_row=2, values_only=True))

def update_materials(data):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["材料"]
    sheet.delete_rows(2, sheet.max_row)
    for row in data:
        sheet.append(row)
    wb.save(EXCEL_FILE)

def get_material_unit_cost(material_name):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["材料"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, amount_str, price_str, tax_str = row
        if name == material_name:
            try:
                # 量と価格、税率を取得して単価を計算
                amount_val = float(''.join(filter(str.isdigit, str(amount_str)))) if amount_str else 0
                price_val = float(''.join(filter(str.isdigit, str(price_str)))) if price_str else 0
                tax_match = re.search(r'([0-9]+(?:\.[0-9]*)?)', str(tax_str))
                tax_val = float(tax_match.group(1)) if tax_match else 0
                if amount_val > 0:
                    # 税込み価格を計算
                    price_incl_tax = price_val * (1 + tax_val / 100)
                    return price_incl_tax / amount_val # 単価を返す
                else:
                    return 0 # 量が0の場合は単価も0
            except Exception:
                return 0 # 計算エラーの場合は単価0
    return 0 # 材料が見つからない場合は単価0

def save_recipe(recipe_name, materials, total_cost):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["レシピ"]
    # ヘッダー行が存在しない場合に追加
    if sheet.max_row == 0 or sheet.cell(row=1, column=1).value is None:
        sheet.append(["レシピ名", "合計原価", "材料名", "量"])
        bold_font = Font(bold=True)
        for col_idx in range(1, 5):
            sheet.cell(row=1, column=col_idx).font = bold_font
    
    # レシピ名と合計原価を最初の行に書き込み
    row_data = [recipe_name, total_cost]
    # 各材料の名前と量を追記
    for mat in materials:
        row_data.append(mat['material_name'])
        row_data.append(mat['amount'])
    sheet.append(row_data)
    wb.save(EXCEL_FILE)

def get_recipes():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["レシピ"]

    # ヘッダー行を除くデータを取得
    # レシピ名、合計原価、その後に材料名と量が続く形式を想定
    recipes = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None: # 空行でないか、レシピ名があるか確認
            recipe_name = row[0]
            total_cost = row[1] if len(row) > 1 and row[1] is not None else "N/A"
            # 材料名のリストと量のリストをペアにして取得 (レシピ名と合計原価の後の要素)
            materials_data = []
            # 材料名、量のリストを2つ組にして取得 (レシピ名と合計原価の後の要素)
            # Excel保存形式が [レシピ名, 合計原価, 材料1名, 量1, 材料2名, 量2, ...] となっている想定
            for i in range(2, len(row), 2):
                if i + 1 < len(row) and row[i] is not None:
                     #materials_data.append({'name': row[i], 'amount': row[i+1]})
                     materials_data.append({'name': row[i], 'amount': row[i+1], 'cost': get_material_unit_cost(row[i])})
            recipes.append({'name': recipe_name, 'total_cost': total_cost, 'materials': materials_data})
    return recipes

def update_recipe(old_recipe_name, new_recipe_name, materials, total_cost):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["レシピ"]
    
    # 古いレシピ名の行を検索して削除
    row_index_to_delete = -1
    for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        if row and row[0].value == old_recipe_name:
            row_index_to_delete = r_idx
            break
    
    if row_index_to_delete != -1:
        sheet.delete_rows(row_index_to_delete, 1)

    # 新しいレシピデータを追加
    row_data = [new_recipe_name, total_cost]
    for mat in materials:
        row_data.append(mat['material_name'])
        row_data.append(mat['amount'])
        # 材料原価もExcelに保存
        # row_data.append(mat['cost'])

    sheet.append(row_data)
    wb.save(EXCEL_FILE)

def delete_recipe(recipe_name):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["レシピ"]
    
    # レシピ名の行を検索して削除
    row_index_to_delete = -1
    for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        if row and row[0].value == recipe_name:
            row_index_to_delete = r_idx
            break
    
    if row_index_to_delete != -1:
        sheet.delete_rows(row_index_to_delete, 1)
        wb.save(EXCEL_FILE)
        return True # 削除成功
    return False # レシピが見つからなかった 